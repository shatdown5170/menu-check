# -*- coding: utf-8 -*-
"""
팀 R&R · 업무량 · 긴급배분 관리 시트 생성 스크립트
- 영역: 기획/PM/운영 팀 (5명 이하)
- 측정: 회당 시간 × 주기 → 월 자동환산
- 가용: 주 32시간 기준
산출물: 팀_RnR_업무관리_시트.xlsx (샘플 + 빈 템플릿 포함)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- 공통 스타일
NAVY   = "1F3864"
BLUE   = "2E5496"
LBLUE  = "D9E1F2"
LLBLUE = "EAF0FA"
GREY   = "808080"
LGREY  = "F2F2F2"
GREEN  = "C6EFCE"; GREEN_T = "006100"
YELLOW = "FFEB9C"; YELLOW_T = "9C6500"
RED    = "FFC7CE"; RED_T = "9C0006"
WHITE  = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def title_font(sz=11, color=WHITE, bold=True):
    return Font(name="맑은 고딕", size=sz, bold=bold, color=color)
def base_font(sz=10, color="000000", bold=False):
    return Font(name="맑은 고딕", size=sz, bold=bold, color=color)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

def fill(color):
    return PatternFill("solid", fgColor=color)

wb = Workbook()

# ================================================================ 참조(hidden)
ref = wb.active
ref.title = "참조"

# 주기 -> 월 발생횟수 (영업일 21일/월, 4.33주/월 기준)
freq_table = [
    ("매일",        21.0),
    ("주3회",       13.0),
    ("주2회",       8.67),
    ("주간(주1회)", 4.33),
    ("격주",        2.17),
    ("월간",        1.0),
    ("분기",        0.33),
    ("반기",        0.17),
    ("연간",        0.083),
    ("수시",        0.0),   # 수시는 '월 발생횟수(직접)' 칸에 직접 입력
]
ref["A1"] = "주기"; ref["B1"] = "월 발생횟수"
for i, (k, v) in enumerate(freq_table, start=2):
    ref[f"A{i}"] = k
    ref[f"B{i}"] = v

# 드롭다운용 리스트들
ref["D1"] = "업무분류"
categories = ["기획", "운영", "데이터/지표", "유관협업", "QA/품질", "관리/지원"]
for i, c in enumerate(categories, start=2):
    ref[f"D{i}"] = c

ref["E1"] = "대체난이도"
diff = ["낮음(누구나 가능)", "중간(인수인계 필요)", "높음(전담 필수)"]
for i, c in enumerate(diff, start=2):
    ref[f"E{i}"] = c

ref["F1"] = "자동화"
auto = ["가능", "일부 가능", "불가"]
for i, c in enumerate(auto, start=2):
    ref[f"F{i}"] = c

ref["G1"] = "우선순위"
prio = ["상", "중", "하"]
for i, c in enumerate(prio, start=2):
    ref[f"G{i}"] = c

ref["H1"] = "팀원"
members = ["김지원(팀장/PM)", "이서연(서비스기획)", "박민준(운영매니저)",
           "최유나(데이터/지표)", "정현우(QA/지원)"]
for i, c in enumerate(members, start=2):
    ref[f"H{i}"] = c

ref["I1"] = "주기리스트"
for i, (k, _v) in enumerate(freq_table, start=2):
    ref[f"I{i}"] = k

ref.sheet_state = "hidden"

# 이름 정의(named ranges) -- 드롭다운/조회 안정화
from openpyxl.workbook.defined_name import DefinedName
wb.defined_names.add(DefinedName("주기표", attr_text="참조!$A$2:$B$11"))
wb.defined_names.add(DefinedName("목록_분류", attr_text="참조!$D$2:$D$7"))
wb.defined_names.add(DefinedName("목록_난이도", attr_text="참조!$E$2:$E$4"))
wb.defined_names.add(DefinedName("목록_자동화", attr_text="참조!$F$2:$F$4"))
wb.defined_names.add(DefinedName("목록_우선순위", attr_text="참조!$G$2:$G$4"))
wb.defined_names.add(DefinedName("목록_팀원", attr_text="참조!$H$2:$H$6"))
wb.defined_names.add(DefinedName("목록_주기", attr_text="참조!$I$2:$I$11"))

# ================================================================ 헬퍼: 업무마스터 시트 빌더
# 컬럼 정의 (헤더, 너비)
COLS = [
    ("No", 5),
    ("업무명", 26),
    ("업무분류", 13),
    ("업무 설명", 34),
    ("주담당", 16),
    ("백업담당", 16),
    ("주기", 12),
    ("회당\n소요시간(h)", 11),
    ("월 발생횟수\n(직접·수시용)", 12),
    ("월 환산\n시간(h)", 11),
    ("필요 역량/스킬", 24),
    ("대체 난이도", 16),
    ("자동화\n가능성", 11),
    ("우선\n순위", 7),
    ("비고", 22),
]
# 컬럼 인덱스 (1-base): G=주기(7) H=회당(8) I=직접(9) J=월환산(10)

def build_master(ws, title_text, data_rows, n_blank=0):
    # ---- 상단 타이틀 영역
    last_col = get_column_letter(len(COLS))
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = title_text
    c.font = title_font(15)
    c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f"A2:{last_col}2")
    g = ws["A2"]
    g.value = ("입력 가이드  |  ① 회당 소요시간 + 주기만 넣으면 '월 환산 시간'이 자동 계산됩니다.  "
               "② '수시' 업무는 '월 발생횟수(직접)' 칸에 한 달 평균 횟수를 직접 입력하세요.  "
               "③ 주담당/백업담당/분류 등은 셀 클릭 시 드롭다운에서 선택.")
    g.font = base_font(9, GREY)
    g.fill = fill(LLBLUE)
    g.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[2].height = 30

    # ---- 헤더 (3행)
    hdr_row = 3
    for j, (name, width) in enumerate(COLS, start=1):
        col = get_column_letter(j)
        ws.column_dimensions[col].width = width
        cell = ws.cell(row=hdr_row, column=j, value=name)
        cell.font = title_font(10)
        cell.fill = fill(BLUE)
        cell.alignment = center
        cell.border = border_all
    ws.row_dimensions[hdr_row].height = 34

    # ---- 데이터 행 작성
    first_data = hdr_row + 1
    r = first_data
    for row in data_rows:
        write_master_row(ws, r, row)
        r += 1
    # ---- 빈 입력 행
    for _ in range(n_blank):
        write_master_row(ws, r, None)
        r += 1
    last_data = r - 1

    # ---- 합계 행
    total_row = last_data + 1
    ws.cell(row=total_row, column=1, value="").fill = fill(LGREY)
    tcell = ws.cell(row=total_row, column=2, value="합계 / 월 총 투입시간")
    tcell.font = base_font(10, NAVY, bold=True)
    tcell.alignment = left
    for j in range(1, len(COLS)+1):
        cc = ws.cell(row=total_row, column=j)
        cc.fill = fill(LBLUE); cc.border = border_all
    sum_cell = ws.cell(row=total_row, column=10,
                       value=f"=SUM(J{first_data}:J{last_data})")
    sum_cell.font = base_font(11, NAVY, bold=True)
    sum_cell.alignment = center
    sum_cell.number_format = "0.0"

    # ---- 데이터 검증(드롭다운) 부착
    add_validations(ws, first_data, last_data)

    # ---- 조건부 서식: 우선순위 '상' 강조, 대체난이도 '높음' 강조
    prio_col = get_column_letter(14)  # N
    diff_col = get_column_letter(12)  # L
    rng_prio = f"{prio_col}{first_data}:{prio_col}{last_data}"
    rng_diff = f"{diff_col}{first_data}:{diff_data_end(last_data)}"
    ws.conditional_formatting.add(
        f"{prio_col}{first_data}:{prio_col}{last_data}",
        CellIsRule(operator="equal", formula=['"상"'],
                   fill=fill(RED), font=Font(name="맑은 고딕", size=10, bold=True, color=RED_T)))
    ws.conditional_formatting.add(
        f"{diff_col}{first_data}:{diff_col}{last_data}",
        CellIsRule(operator="equal", formula=['"높음(전담 필수)"'],
                   fill=fill(YELLOW), font=Font(name="맑은 고딕", size=10, color=YELLOW_T)))

    # 틀 고정 (헤더까지)
    ws.freeze_panes = f"A{first_data}"
    ws.sheet_view.showGridLines = False
    return first_data, last_data

def diff_data_end(last_data):
    return last_data

def write_master_row(ws, r, row):
    """row = dict or None(빈행)"""
    for j in range(1, len(COLS)+1):
        cell = ws.cell(row=r, column=j)
        cell.border = border_all
        cell.font = base_font(10)
        if j in (1, 7, 8, 9, 10, 13, 14):
            cell.alignment = center
        else:
            cell.alignment = left
        if r % 2 == 0:
            cell.fill = fill(WHITE)
        else:
            cell.fill = fill(LGREY)
    # No
    ws.cell(row=r, column=1, value=f"=IF($B{r}=\"\",\"\",ROW()-3)")
    # 월 환산 시간(J) 자동 계산 공식 (모든 행에 부여)
    # 유효 월횟수 = 직접입력(I)이 있으면 그 값, 없으면 주기(G)로 VLOOKUP
    jf = (f'=IF($B{r}="","",'
          f'$H{r}*IF($I{r}<>"",$I{r},IFERROR(VLOOKUP($G{r},주기표,2,FALSE),0)))')
    jc = ws.cell(row=r, column=10, value=jf)
    jc.number_format = "0.0"
    # 회당시간 숫자 포맷
    ws.cell(row=r, column=8).number_format = "0.0"
    ws.cell(row=r, column=9).number_format = "0.0"
    if row is None:
        return
    ws.cell(row=r, column=2, value=row["업무명"])
    ws.cell(row=r, column=3, value=row["분류"])
    ws.cell(row=r, column=4, value=row["설명"])
    ws.cell(row=r, column=5, value=row["주담당"])
    ws.cell(row=r, column=6, value=row["백업"])
    ws.cell(row=r, column=7, value=row["주기"])
    ws.cell(row=r, column=8, value=row["회당"])
    if row.get("월횟수직접") is not None:
        ws.cell(row=r, column=9, value=row["월횟수직접"])
    ws.cell(row=r, column=11, value=row["역량"])
    ws.cell(row=r, column=12, value=row["난이도"])
    ws.cell(row=r, column=13, value=row["자동화"])
    ws.cell(row=r, column=14, value=row["우선순위"])
    ws.cell(row=r, column=15, value=row.get("비고", ""))

def add_validations(ws, first_data, last_data):
    def dv(formula):
        d = DataValidation(type="list", formula1=formula, allow_blank=True)
        d.error = "목록에서 선택하세요"; d.errorTitle = "입력값 확인"
        return d
    specs = [
        ("C", "=목록_분류"),
        ("E", "=목록_팀원"),
        ("F", "=목록_팀원"),
        ("G", "=목록_주기"),
        ("L", "=목록_난이도"),
        ("M", "=목록_자동화"),
        ("N", "=목록_우선순위"),
    ]
    for col, f in specs:
        d = dv(f)
        ws.add_data_validation(d)
        d.add(f"{col}{first_data}:{col}{last_data}")

# ================================================================ 샘플 데이터
SAMPLE = [
    dict(업무명="주간 팀 회의 준비·진행", 분류="운영", 설명="아젠다 취합, 회의록 작성·공유",
         주담당="김지원(팀장/PM)", 백업="이서연(서비스기획)", 주기="주간(주1회)", 회당=1.5,
         역량="퍼실리테이션, 문서화", 난이도="낮음(누구나 가능)", 자동화="일부 가능", 우선순위="중",
         비고="회의록 템플릿화로 단축 여지"),
    dict(업무명="월간 사업보고서 작성", 분류="기획", 설명="실적·지표 종합, 경영진 보고자료 작성",
         주담당="김지원(팀장/PM)", 백업="최유나(데이터/지표)", 주기="월간", 회당=6,
         역량="데이터 해석, 스토리텔링", 난이도="높음(전담 필수)", 자동화="일부 가능", 우선순위="상",
         비고="지표는 최유나 협조"),
    dict(업무명="분기 OKR 수립·리뷰", 분류="기획", 설명="목표 설정, 핵심지표 정의, 회고",
         주담당="김지원(팀장/PM)", 백업="이서연(서비스기획)", 주기="분기", 회당=8,
         역량="전략기획", 난이도="높음(전담 필수)", 자동화="불가", 우선순위="상", 비고=""),
    dict(업무명="유관부서 협업 미팅", 분류="유관협업", 설명="개발·디자인·마케팅 정례 싱크",
         주담당="김지원(팀장/PM)", 백업="박민준(운영매니저)", 주기="주간(주1회)", 회당=2,
         역량="커뮤니케이션, 조율", 난이도="중간(인수인계 필요)", 자동화="불가", 우선순위="중", 비고=""),
    dict(업무명="서비스 기능 기획서 작성", 분류="기획", 설명="요구사항 정의, 기획서·스펙 문서화",
         주담당="이서연(서비스기획)", 백업="김지원(팀장/PM)", 주기="격주", 회당=8,
         역량="서비스기획, UX", 난이도="높음(전담 필수)", 자동화="불가", 우선순위="상", 비고=""),
    dict(업무명="사용자 피드백 정리·분석", 분류="기획", 설명="VOC·설문 취합, 인사이트 도출",
         주담당="이서연(서비스기획)", 백업="박민준(운영매니저)", 주기="주간(주1회)", 회당=3,
         역량="리서치, 정성분석", 난이도="중간(인수인계 필요)", 자동화="일부 가능", 우선순위="중", 비고=""),
    dict(업무명="화면 정책·플로우 정의", 분류="기획", 설명="정책 정의, 화면 플로우/예외처리 정리",
         주담당="이서연(서비스기획)", 백업="정현우(QA/지원)", 주기="격주", 회당=5,
         역량="기획, 논리설계", 난이도="높음(전담 필수)", 자동화="불가", 우선순위="중", 비고=""),
    dict(업무명="릴리즈 노트 작성", 분류="운영", 설명="배포 내역 정리, 사내 공지",
         주담당="이서연(서비스기획)", 백업="정현우(QA/지원)", 주기="월간", 회당=2,
         역량="문서화", 난이도="낮음(누구나 가능)", 자동화="가능", 우선순위="하", 비고="자동화 1순위 후보"),
    dict(업무명="일일 운영 모니터링", 분류="운영", 설명="대시보드·알람 점검, 이상징후 확인",
         주담당="박민준(운영매니저)", 백업="최유나(데이터/지표)", 주기="매일", 회당=0.5,
         역량="운영, 모니터링", 난이도="중간(인수인계 필요)", 자동화="일부 가능", 우선순위="상", 비고=""),
    dict(업무명="CS 이슈 대응·에스컬레이션", 분류="운영", 설명="문의 분류, 처리·유관부서 전달",
         주담당="박민준(운영매니저)", 백업="정현우(QA/지원)", 주기="매일", 회당=1,
         역량="CS, 문제해결", 난이도="중간(인수인계 필요)", 자동화="불가", 우선순위="상", 비고="피크시 변동 큼"),
    dict(업무명="운영 프로세스 문서 업데이트", 분류="관리/지원", 설명="매뉴얼·플레이북 현행화",
         주담당="박민준(운영매니저)", 백업="이서연(서비스기획)", 주기="월간", 회당=3,
         역량="문서화, 프로세스설계", 난이도="낮음(누구나 가능)", 자동화="불가", 우선순위="하", 비고=""),
    dict(업무명="외부 벤더 커뮤니케이션", 분류="유관협업", 설명="제휴사·외주 일정·이슈 조율",
         주담당="박민준(운영매니저)", 백업="김지원(팀장/PM)", 주기="주간(주1회)", 회당=1.5,
         역량="협상, 커뮤니케이션", 난이도="중간(인수인계 필요)", 자동화="불가", 우선순위="중", 비고=""),
    dict(업무명="주간 지표 대시보드 업데이트", 분류="데이터/지표", 설명="핵심지표 집계·갱신, 코멘트",
         주담당="최유나(데이터/지표)", 백업="박민준(운영매니저)", 주기="주간(주1회)", 회당=2,
         역량="SQL, 데이터시각화", 난이도="중간(인수인계 필요)", 자동화="가능", 우선순위="중", 비고="쿼리 자동화 가능"),
    dict(업무명="월간 성과 데이터 분석 리포트", 분류="데이터/지표", 설명="성과 심층분석, 개선 제언",
         주담당="최유나(데이터/지표)", 백업="김지원(팀장/PM)", 주기="월간", 회당=8,
         역량="데이터분석, 통계", 난이도="높음(전담 필수)", 자동화="일부 가능", 우선순위="상", 비고=""),
    dict(업무명="A/B 테스트 설계·분석", 분류="데이터/지표", 설명="실험 설계, 결과 해석·의사결정 지원",
         주담당="최유나(데이터/지표)", 백업="이서연(서비스기획)", 주기="격주", 회당=5,
         역량="실험설계, 통계", 난이도="높음(전담 필수)", 자동화="불가", 우선순위="중", 비고=""),
    dict(업무명="데이터 추출 Ad-hoc 요청 대응", 분류="데이터/지표", 설명="비정기 데이터 추출·가공 요청 처리",
         주담당="최유나(데이터/지표)", 백업="박민준(운영매니저)", 주기="수시", 회당=1, 월횟수직접=6,
         역량="SQL", 난이도="중간(인수인계 필요)", 자동화="일부 가능", 우선순위="중",
         비고="월 평균 6건 가정(직접입력 예시)"),
    dict(업무명="QA 테스트 케이스 작성", 분류="QA/품질", 설명="기능별 테스트 시나리오 설계",
         주담당="정현우(QA/지원)", 백업="이서연(서비스기획)", 주기="격주", 회당=4,
         역량="QA, 테스트설계", 난이도="중간(인수인계 필요)", 자동화="일부 가능", 우선순위="중", 비고=""),
    dict(업무명="배포 전 QA 검증", 분류="QA/품질", 설명="릴리즈 전 회귀·시나리오 검증",
         주담당="정현우(QA/지원)", 백업="박민준(운영매니저)", 주기="격주", 회당=3,
         역량="QA, 꼼꼼함", 난이도="중간(인수인계 필요)", 자동화="일부 가능", 우선순위="상", 비고="배포 주기 연동"),
    dict(업무명="버그 트래킹·리포팅", 분류="QA/품질", 설명="이슈 등록·우선순위화, 처리현황 공유",
         주담당="정현우(QA/지원)", 백업="박민준(운영매니저)", 주기="주간(주1회)", 회당=2,
         역량="이슈관리", 난이도="낮음(누구나 가능)", 자동화="일부 가능", 우선순위="중", 비고=""),
    dict(업무명="내부 위키·문서 정리", 분류="관리/지원", 설명="팀 지식베이스 구조화·정리",
         주담당="정현우(QA/지원)", 백업="이서연(서비스기획)", 주기="월간", 회당=3,
         역량="문서화, 정리", 난이도="낮음(누구나 가능)", 자동화="불가", 우선순위="하", 비고=""),
    # ---- 추가 업무 (현실적 워크로드 반영) ----
    dict(업무명="경영진 주간 실적 보고", 분류="기획", 설명="주요 지표·이슈 요약 보고",
         주담당="김지원(팀장/PM)", 백업="최유나(데이터/지표)", 주기="주간(주1회)", 회당=2,
         역량="요약, 보고", 난이도="중간(인수인계 필요)", 자동화="일부 가능", 우선순위="상", 비고=""),
    dict(업무명="채용·면접 진행", 분류="관리/지원", 설명="서류검토·면접·피드백",
         주담당="김지원(팀장/PM)", 백업="이서연(서비스기획)", 주기="수시", 회당=2, 월횟수직접=3,
         역량="평가, 커뮤니케이션", 난이도="중간(인수인계 필요)", 자동화="불가", 우선순위="중",
         비고="채용시즌 변동 큼"),
    dict(업무명="디자인 QA·리뷰", 분류="기획", 설명="산출물 디자인 검수·피드백",
         주담당="이서연(서비스기획)", 백업="정현우(QA/지원)", 주기="주간(주1회)", 회당=2,
         역량="UX, 디테일", 난이도="중간(인수인계 필요)", 자동화="불가", 우선순위="중", 비고=""),
    dict(업무명="스프린트 백로그 관리", 분류="기획", 설명="이슈 정리·우선순위화·그루밍",
         주담당="이서연(서비스기획)", 백업="김지원(팀장/PM)", 주기="주간(주1회)", 회당=1.5,
         역량="기획, 일정관리", 난이도="중간(인수인계 필요)", 자동화="일부 가능", 우선순위="중", 비고=""),
    dict(업무명="경쟁사·시장 리서치", 분류="기획", 설명="동향 모니터링·벤치마킹 정리",
         주담당="이서연(서비스기획)", 백업="최유나(데이터/지표)", 주기="월간", 회당=4,
         역량="리서치", 난이도="낮음(누구나 가능)", 자동화="불가", 우선순위="하", 비고=""),
    dict(업무명="주간 운영 현황 리포트", 분류="데이터/지표", 설명="운영 지표 집계·공유",
         주담당="박민준(운영매니저)", 백업="최유나(데이터/지표)", 주기="주간(주1회)", 회당=2,
         역량="데이터 정리", 난이도="중간(인수인계 필요)", 자동화="가능", 우선순위="중", 비고="자동화 후보"),
    dict(업무명="정산·매출 데이터 점검", 분류="운영", 설명="정산 데이터 검증·이상 확인",
         주담당="박민준(운영매니저)", 백업="최유나(데이터/지표)", 주기="주간(주1회)", 회당=1.5,
         역량="꼼꼼함, 숫자감각", 난이도="중간(인수인계 필요)", 자동화="일부 가능", 우선순위="상", 비고=""),
    dict(업무명="신규 입점·온보딩 처리", 분류="운영", 설명="신규 건 검수·등록·안내",
         주담당="박민준(운영매니저)", 백업="정현우(QA/지원)", 주기="수시", 회당=0.75, 월횟수직접=8,
         역량="운영, 꼼꼼함", 난이도="낮음(누구나 가능)", 자동화="일부 가능", 우선순위="중", 비고=""),
    dict(업무명="장애·인시던트 대응", 분류="운영", 설명="긴급 이슈 처리·사후 리포트",
         주담당="박민준(운영매니저)", 백업="정현우(QA/지원)", 주기="수시", 회당=2, 월횟수직접=4,
         역량="문제해결, 침착함", 난이도="높음(전담 필수)", 자동화="불가", 우선순위="상",
         비고="발생 예측 어려움"),
    dict(업무명="일일 정산 마감 확인", 분류="운영", 설명="마감 배치·수치 확인",
         주담당="박민준(운영매니저)", 백업="최유나(데이터/지표)", 주기="매일", 회당=0.5,
         역량="운영", 난이도="중간(인수인계 필요)", 자동화="가능", 우선순위="상", 비고="자동화 후보"),
    dict(업무명="고객 문의 2차 대응", 분류="운영", 설명="에스컬레이션 건 심층 처리",
         주담당="박민준(운영매니저)", 백업="정현우(QA/지원)", 주기="매일", 회당=0.5,
         역량="CS, 문제해결", 난이도="중간(인수인계 필요)", 자동화="불가", 우선순위="중", 비고=""),
    dict(업무명="주간 운영 회의 주재", 분류="운영", 설명="운영 이슈 점검 회의 진행",
         주담당="박민준(운영매니저)", 백업="김지원(팀장/PM)", 주기="주간(주1회)", 회당=1.5,
         역량="퍼실리테이션", 난이도="낮음(누구나 가능)", 자동화="불가", 우선순위="중", 비고=""),
    dict(업무명="데이터 거버넌스·지표 정의", 분류="데이터/지표", 설명="지표 표준·정의 관리",
         주담당="최유나(데이터/지표)", 백업="이서연(서비스기획)", 주기="월간", 회당=4,
         역량="데이터 모델링", 난이도="높음(전담 필수)", 자동화="불가", 우선순위="중", 비고=""),
    dict(업무명="정기 코호트 분석", 분류="데이터/지표", 설명="리텐션·코호트 추이 분석",
         주담당="최유나(데이터/지표)", 백업="김지원(팀장/PM)", 주기="격주", 회당=4,
         역량="데이터분석", 난이도="높음(전담 필수)", 자동화="일부 가능", 우선순위="중", 비고=""),
    dict(업무명="일일 스모크 테스트", 분류="QA/품질", 설명="핵심 플로우 일일 점검",
         주담당="정현우(QA/지원)", 백업="박민준(운영매니저)", 주기="매일", 회당=0.5,
         역량="QA", 난이도="낮음(누구나 가능)", 자동화="가능", 우선순위="중", 비고="자동화 후보"),
    dict(업무명="QA 자동화 스크립트 유지보수", 분류="QA/품질", 설명="테스트 자동화 코드 관리",
         주담당="정현우(QA/지원)", 백업="최유나(데이터/지표)", 주기="주간(주1회)", 회당=2,
         역량="테스트 자동화, 코드", 난이도="높음(전담 필수)", 자동화="일부 가능", 우선순위="중", 비고=""),
]

# ================================================================ ① 업무마스터(샘플)
ws_master = wb.create_sheet("① 업무마스터(R&R)")
f0, fL = build_master(ws_master, "① 업무 마스터  ·  R&R 정의표  (샘플 데이터)", SAMPLE, n_blank=6)
MASTER = "'① 업무마스터(R&R)'"

# ================================================================ ② 팀원별 부하(대시보드)
ws_dash = wb.create_sheet("② 팀원별 부하")
ws_dash.sheet_view.showGridLines = False
WEEKLY_CAP = 32  # 주당 가용시간
MONTH_FACTOR = 4.33

dcols = [
    ("팀원", 22), ("직무/역할", 16), ("주당\n가용(h)", 10), ("월 가용\n(h)", 10),
    ("월 투입(h)\n(주담당)", 12), ("담당\n업무수", 9), ("부하율", 10),
    ("여력(h)\n(월)", 11), ("상태", 12), ("백업\n지정수", 9),
]
last_col_d = get_column_letter(len(dcols))
ws_dash.merge_cells(f"A1:{last_col_d}1")
t = ws_dash["A1"]; t.value = "② 팀원별 업무 부하 대시보드  (자동 계산)"
t.font = title_font(15); t.fill = fill(NAVY)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_dash.row_dimensions[1].height = 34

ws_dash.merge_cells(f"A2:{last_col_d}2")
gd = ws_dash["A2"]
gd.value = (f"기준: 주당 가용 {WEEKLY_CAP}시간(회의·휴식 제외 실작업) × {MONTH_FACTOR}주 = 월 가용시간.  "
            "투입시간은 ① 업무마스터의 '주담당' 기준으로 자동 합산됩니다.  "
            "부하율 70%↑ 과부하 / 40~70% 적정 / 40%↓ 여유. (정기 업무 기준 — 나머지는 회의·돌발업무 몫)")
gd.font = base_font(9, GREY); gd.fill = fill(LLBLUE)
gd.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
ws_dash.row_dimensions[2].height = 30

hr = 3
for j, (name, width) in enumerate(dcols, start=1):
    col = get_column_letter(j)
    ws_dash.column_dimensions[col].width = width
    c = ws_dash.cell(row=hr, column=j, value=name)
    c.font = title_font(10); c.fill = fill(BLUE); c.alignment = center; c.border = border_all
ws_dash.row_dimensions[hr].height = 34

roles = ["팀장 / PM", "서비스 기획", "운영 매니저", "데이터 / 지표", "QA / 지원"]
dfirst = hr + 1
for idx, (m, role) in enumerate(zip(members, roles)):
    r = dfirst + idx
    ws_dash.cell(row=r, column=1, value=m).alignment = left
    ws_dash.cell(row=r, column=2, value=role).alignment = center
    ws_dash.cell(row=r, column=3, value=WEEKLY_CAP).alignment = center
    ws_dash.cell(row=r, column=4, value=f"=C{r}*{MONTH_FACTOR}").alignment = center
    # 월 투입 = SUMIF(주담당=팀원, 월환산시간)
    ws_dash.cell(row=r, column=5,
        value=f"=SUMIF({MASTER}!$E${f0}:$E${fL},$A{r},{MASTER}!$J${f0}:$J${fL})").alignment = center
    ws_dash.cell(row=r, column=6,
        value=f"=COUNTIF({MASTER}!$E${f0}:$E${fL},$A{r})").alignment = center
    ws_dash.cell(row=r, column=7, value=f"=IFERROR(E{r}/D{r},0)").alignment = center
    ws_dash.cell(row=r, column=8, value=f"=D{r}-E{r}").alignment = center
    ws_dash.cell(row=r, column=9,
        value=f'=IF(G{r}>=0.7,"⚠ 과부하",IF(G{r}>=0.4,"적정","🟢 여유"))').alignment = center
    ws_dash.cell(row=r, column=10,
        value=f"=COUNTIF({MASTER}!$F${f0}:$F${fL},$A{r})").alignment = center
    for j in range(1, len(dcols)+1):
        cc = ws_dash.cell(row=r, column=j)
        cc.border = border_all; cc.font = base_font(10)
        if cc.fill.fgColor.rgb in (None, "00000000"):
            cc.fill = fill(WHITE if idx % 2 == 0 else LGREY)
    ws_dash.cell(row=r, column=4).number_format = "0.0"
    ws_dash.cell(row=r, column=5).number_format = "0.0"
    ws_dash.cell(row=r, column=7).number_format = "0%"
    ws_dash.cell(row=r, column=8).number_format = "0.0"
dlast = dfirst + len(members) - 1

# 합계/평균 행
tr = dlast + 1
ws_dash.cell(row=tr, column=1, value="팀 합계 / 평균").font = base_font(10, NAVY, bold=True)
ws_dash.cell(row=tr, column=1).alignment = left
ws_dash.cell(row=tr, column=4, value=f"=SUM(D{dfirst}:D{dlast})").number_format = "0.0"
ws_dash.cell(row=tr, column=5, value=f"=SUM(E{dfirst}:E{dlast})").number_format = "0.0"
ws_dash.cell(row=tr, column=6, value=f"=SUM(F{dfirst}:F{dlast})")
ws_dash.cell(row=tr, column=7, value=f"=IFERROR(E{tr}/D{tr},0)").number_format = "0%"
ws_dash.cell(row=tr, column=8, value=f"=SUM(H{dfirst}:H{dlast})").number_format = "0.0"
for j in range(1, len(dcols)+1):
    cc = ws_dash.cell(row=tr, column=j)
    cc.fill = fill(LBLUE); cc.border = border_all
    if cc.font.bold is not True:
        cc.font = base_font(10, NAVY, bold=True)
    cc.alignment = center
ws_dash.cell(row=tr, column=1).alignment = left

# 조건부 서식: 부하율 컬러스케일 + 상태 색상
ws_dash.conditional_formatting.add(
    f"G{dfirst}:G{dlast}",
    ColorScaleRule(start_type="num", start_value=0, start_color=GREEN,
                   mid_type="num", mid_value=0.55, mid_color=YELLOW,
                   end_type="num", end_value=1.0, end_color=RED))
ws_dash.conditional_formatting.add(
    f"I{dfirst}:I{dlast}",
    FormulaRule(formula=[f'ISNUMBER(SEARCH("과부하",I{dfirst}))'],
                fill=fill(RED), font=Font(name="맑은 고딕", size=10, bold=True, color=RED_T)))
ws_dash.conditional_formatting.add(
    f"I{dfirst}:I{dlast}",
    FormulaRule(formula=[f'ISNUMBER(SEARCH("여유",I{dfirst}))'],
                fill=fill(GREEN), font=Font(name="맑은 고딕", size=10, bold=True, color=GREEN_T)))
ws_dash.freeze_panes = f"A{dfirst}"

# ================================================================ ③ 긴급배분 가이드
ws_urg = wb.create_sheet("③ 긴급배분 가이드")
ws_urg.sheet_view.showGridLines = False
for col, w in zip("ABCDEFG", [22, 14, 12, 12, 12, 16, 30]):
    ws_urg.column_dimensions[col].width = w

ws_urg.merge_cells("A1:G1")
t = ws_urg["A1"]; t.value = "③ 긴급 업무 배분 가이드"
t.font = title_font(15); t.fill = fill(NAVY)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_urg.row_dimensions[1].height = 34

ws_urg.merge_cells("A2:G2")
gu = ws_urg["A2"]
gu.value = ("갑작스런 업무 발생 시 ① 필요 역량이 맞는 사람을 먼저 추리고 → ② 아래 표에서 '여력(h)'이 큰 사람에게 배정하세요.  "
            "B3 셀에 예상 소요시간(월,h)을 입력하면 누가 수용 가능한지 자동 표시됩니다.")
gu.font = base_font(9, GREY); gu.fill = fill(LLBLUE)
gu.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
ws_urg.row_dimensions[2].height = 32

# 입력 셀
ws_urg["A3"] = "▶ 긴급업무 예상 소요시간 (월, h):"
ws_urg["A3"].font = base_font(10, NAVY, bold=True)
inp = ws_urg["B3"]; inp.value = 45
inp.fill = fill("FFF2CC"); inp.font = base_font(11, bold=True); inp.alignment = center
inp.border = Border(left=Side(style="medium", color=YELLOW_T),
                    right=Side(style="medium", color=YELLOW_T),
                    top=Side(style="medium", color=YELLOW_T),
                    bottom=Side(style="medium", color=YELLOW_T))
ws_urg["C3"] = "← 노란 칸에 입력"
ws_urg["C3"].font = base_font(9, GREY); ws_urg.merge_cells("C3:D3")

# 헤더
uhdr = ["팀원", "현재 부하율", "여력(h,월)", "상태", "수용 가능?", "여유 후 부하율", "비고"]
hr2 = 5
for j, name in enumerate(uhdr, start=1):
    c = ws_urg.cell(row=hr2, column=j, value=name)
    c.font = title_font(10); c.fill = fill(BLUE); c.alignment = center; c.border = border_all
ws_urg.row_dimensions[hr2].height = 30

ufirst = hr2 + 1
for idx in range(len(members)):
    dr = dfirst + idx  # dashboard row
    r = ufirst + idx
    ws_urg.cell(row=r, column=1, value=f"='② 팀원별 부하'!A{dr}").alignment = left
    ws_urg.cell(row=r, column=2, value=f"='② 팀원별 부하'!G{dr}").alignment = center
    ws_urg.cell(row=r, column=2).number_format = "0%"
    ws_urg.cell(row=r, column=3, value=f"='② 팀원별 부하'!H{dr}").alignment = center
    ws_urg.cell(row=r, column=3).number_format = "0.0"
    ws_urg.cell(row=r, column=4, value=f"='② 팀원별 부하'!I{dr}").alignment = center
    ws_urg.cell(row=r, column=5,
        value=f'=IF(C{r}>=$B$3,"✅ 가능","❌ 어려움")').alignment = center
    ws_urg.cell(row=r, column=6,
        value=f"=IFERROR(('② 팀원별 부하'!E{dr}+$B$3)/'② 팀원별 부하'!D{dr},0)").alignment = center
    ws_urg.cell(row=r, column=6).number_format = "0%"
    ws_urg.cell(row=r, column=7, value="").alignment = left
    for j in range(1, 8):
        cc = ws_urg.cell(row=r, column=j)
        cc.border = border_all; cc.font = base_font(10)
        cc.fill = fill(WHITE if idx % 2 == 0 else LGREY)
ulast = ufirst + len(members) - 1

ws_urg.conditional_formatting.add(
    f"E{ufirst}:E{ulast}",
    FormulaRule(formula=[f'ISNUMBER(SEARCH("가능",E{ufirst}))'],
                fill=fill(GREEN), font=Font(name="맑은 고딕", size=10, bold=True, color=GREEN_T)))
ws_urg.conditional_formatting.add(
    f"E{ufirst}:E{ulast}",
    FormulaRule(formula=[f'ISNUMBER(SEARCH("어려움",E{ufirst}))'],
                fill=fill(RED), font=Font(name="맑은 고딕", size=10, color=RED_T)))
ws_urg.conditional_formatting.add(
    f"F{ufirst}:F{ulast}",
    CellIsRule(operator="greaterThan", formula=["1"],
               fill=fill(RED), font=Font(name="맑은 고딕", size=10, bold=True, color=RED_T)))

# 추가 안내 박스
note_r = ulast + 2
ws_urg.merge_cells(f"A{note_r}:G{note_r}")
nb = ws_urg.cell(row=note_r, column=1)
nb.value = "💡 활용 팁"
nb.font = base_font(11, NAVY, bold=True)
tips = [
    "• 역량 매칭: 먼저 '① 업무마스터'의 '필요 역량/스킬'·'백업담당'을 보고 후보를 좁히세요.",
    "• 대체 난이도가 '높음(전담 필수)'인 업무는 긴급 이관이 어렵습니다 — 백업담당 사전 양성이 핵심.",
    "• '여유 후 부하율'이 100%를 넘으면(빨강) 그 사람도 과부하가 되므로 분할 배정을 검토하세요.",
    "• 반복적으로 과부하가 잡히면 '자동화 가능' 업무부터 줄여 구조적으로 여력을 확보하세요.",
]
for i, tip in enumerate(tips, start=note_r+1):
    ws_urg.merge_cells(f"A{i}:G{i}")
    c = ws_urg.cell(row=i, column=1, value=tip)
    c.font = base_font(10); c.alignment = left_top
    ws_urg.row_dimensions[i].height = 20

# ================================================================ 빈 템플릿
ws_blank = wb.create_sheet("빈 템플릿")
build_master(ws_blank, "업무 마스터  ·  R&R 정의표  (빈 템플릿 — 우리 팀 내용으로 채우세요)", [], n_blank=30)

# ================================================================ 사용안내 (맨 앞)
ws_guide = wb.create_sheet("📖 사용안내")
ws_guide.sheet_view.showGridLines = False
for col, w in zip("ABCDEFG", [3, 22, 60, 12, 12, 12, 12]):
    ws_guide.column_dimensions[col].width = w

ws_guide.merge_cells("B2:F2")
t = ws_guide["B2"]; t.value = "팀 R&R · 업무량 · 긴급배분 관리 시트"
t.font = title_font(18, NAVY); t.fill = fill(LLBLUE)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_guide.row_dimensions[2].height = 40

ws_guide.merge_cells("B3:F3")
st = ws_guide["B3"]
st.value = "기획/PM/운영 팀용 · 주당 가용 32시간 기준 · 회당시간×주기 자동 월환산"
st.font = base_font(10, GREY); st.alignment = Alignment(horizontal="left", indent=1)

def guide_section(row, title, lines):
    ws_guide.merge_cells(f"B{row}:F{row}")
    c = ws_guide.cell(row=row, column=2, value=title)
    c.font = title_font(12, WHITE); c.fill = fill(BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_guide.row_dimensions[row].height = 26
    r = row + 1
    for ln in lines:
        ws_guide.merge_cells(f"B{r}:F{r}")
        cc = ws_guide.cell(row=r, column=2, value=ln)
        cc.font = base_font(10); cc.alignment = left_top
        ws_guide.row_dimensions[r].height = 20
        r += 1
    return r + 1

r = 5
r = guide_section(r, "이 파일의 구성", [
    "📑 ① 업무마스터(R&R) : 모든 업무를 한 줄씩 기록하는 핵심 시트 (샘플 20개 포함). 누가 주담당/백업인지, 얼마나 걸리는지, 누가 꼭 해야 하는지 정리.",
    "📊 ② 팀원별 부하 : 위 데이터를 사람별로 자동 합산 — 누가 과부하이고 누가 여유 있는지 한눈에.",
    "🚨 ③ 긴급배분 가이드 : 갑작스런 업무가 생겼을 때 '누가 받을 수 있는지'를 예상 소요시간 입력만으로 확인.",
    "📝 빈 템플릿 : 샘플을 지운 깨끗한 양식. 우리 팀 실제 내용으로 채우면 됩니다.",
])
r = guide_section(r, "작성 순서 (3단계)", [
    "1단계  '빈 템플릿' 또는 '① 업무마스터'에 팀의 모든 업무를 한 줄씩 적습니다. (먼저 떠오르는 대로, 빠짐없이)",
    "2단계  각 업무의 '회당 소요시간'과 '주기'를 입력 → '월 환산 시간'이 자동 계산됩니다. '수시' 업무는 '월 발생횟수(직접)'에 한 달 평균 횟수 입력.",
    "3단계  '주담당/백업담당/대체난이도/자동화/우선순위'를 드롭다운에서 선택. → ②·③ 시트는 자동으로 채워집니다.",
])
r = guide_section(r, "각 항목의 의미", [
    "• 주담당 / 백업담당 : 평소 담당자와, 부재·과부하 시 대신할 사람. (백업이 비어있으면 그 업무는 '버스 팩터' 위험!)",
    "• 주기 : 매일·주간·격주·월간·분기 등. 월 발생횟수로 환산되어 업무량 계산의 기준이 됩니다.",
    "• 대체 난이도 : '낮음=누구나' / '중간=인수인계 필요' / '높음=전담 필수'. 긴급 이관 가능성을 가늠하는 핵심 지표.",
    "• 자동화 가능성 : '가능/일부/불가'. 구조적으로 업무량을 줄일 후보를 찾는 데 사용.",
    "• 우선순위 : 상/중/하. 여력이 부족할 때 무엇을 먼저 지킬지 판단.",
])
r = guide_section(r, "부하율 읽는 법", [
    "부하율 = 월 투입시간 ÷ 월 가용시간(주32h×4.33).   70% 이상 = ⚠ 과부하 / 40~70% = 적정 / 40% 미만 = 🟢 여유",
    "※ '정기적으로 추적되는 업무'만 합산한 값입니다. 나머지 시간은 회의·소통·돌발업무·집중작업의 몫이며, 그 자체가 긴급 업무를 받을 '여력'입니다.",
    "※ 가용시간 기준(주당 32h)은 '② 팀원별 부하' 시트의 '주당 가용(h)' 칸에서 사람마다 조정할 수 있습니다.",
])
ws_guide.merge_cells(f"B{r}:F{r}")
fin = ws_guide.cell(row=r, column=2,
    value="문의·수정 필요 시 언제든 말씀해 주세요. 컬럼 추가/삭제, 부서별 분리, 구글시트 변환 모두 가능합니다.")
fin.font = base_font(10, GREY, bold=True); fin.alignment = left

# 시트 순서 정리: 사용안내 -> 마스터 -> 부하 -> 긴급 -> 빈템플릿 -> 참조(hidden)
order = ["📖 사용안내", "① 업무마스터(R&R)", "② 팀원별 부하", "③ 긴급배분 가이드", "빈 템플릿", "참조"]
wb._sheets.sort(key=lambda s: order.index(s.title))
wb.active = wb.sheetnames.index("📖 사용안내")

# 탭 색상
colors = {"📖 사용안내": NAVY, "① 업무마스터(R&R)": BLUE, "② 팀원별 부하": "548235",
          "③ 긴급배분 가이드": "C55A11", "빈 템플릿": "7F7F7F"}
for name, col in colors.items():
    wb[name].sheet_properties.tabColor = col

out = "/home/user/menu-check/팀_RnR_업무관리_시트.xlsx"
wb.save(out)
print("saved:", out)
